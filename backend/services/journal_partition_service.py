"""CAS/JCR journal partition import and lookup helpers."""
from __future__ import annotations

import difflib
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.config import BASE_DIR
from backend.models import JournalPartition


ALIAS_PAIRS = {
    "Brief in bioinformatics": "BRIEFINGS IN BIOINFORMATICS",
    "Cluster Comput": "Cluster Computing-The Journal of Networks Software Tools and Applications",
    "Future Generation Computer Systems": "Future Generation Computer Systems-The International Journal of eScience",
    "Applied Sciences": "Applied Sciences-Basel",
    "IEEE/CAA Journal of Automatica Sinica": "IEEE-CAA Journal of Automatica Sinica",
}

DOI_JOURNAL_CODES = {
    "j.engappai": "ENGINEERING APPLICATIONS OF ARTIFICIAL INTELLIGENCE",
    "j.patcog": "PATTERN RECOGNITION",
    "j.knosys": "KNOWLEDGE-BASED SYSTEMS",
    "j.eswa": "EXPERT SYSTEMS WITH APPLICATIONS",
    "j.measurement": "MEASUREMENT",
}

PARTITION_LABELS = {
    1: "一区",
    2: "二区",
    3: "三区",
    4: "四区",
}


def normalize_journal_name(value: str | None) -> str:
    text = str(value or "").replace("&", " and ")
    for ch in ["‐", "‑", "–", "—", "－", "/", "-"]:
        text = text.replace(ch, " ")
    text = unicodedata.normalize("NFKD", text)
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).lower()
    return re.sub(r"\s+", " ", text).strip()


def is_yes(value: Any) -> bool:
    return str(value or "").strip() in {"是", "ÊÇ", "Yes", "YES", "Y", "1", "True", "true"}


def source_type(category: str | None) -> str:
    value = str(category or "").upper()
    if "SSCI" in value:
        return "SSCI"
    if "SCIE" in value:
        return "SCIE"
    if "ESCI" in value:
        return "ESCI"
    return "未知"


def partition_label(value: Any) -> str:
    try:
        code = int(str(value).strip())
    except (TypeError, ValueError):
        match = re.search(r"[一二三四1234]", str(value or ""))
        if not match:
            return "未收录"
        token = match.group(0)
        code = int(token) if token.isdigit() else {"一": 1, "二": 2, "三": 3, "四": 4}.get(token, 0)
    return PARTITION_LABELS.get(code, "未收录")


def partition_code(value: Any) -> int | None:
    label = partition_label(value)
    return {"一区": 1, "二区": 2, "三区": 3, "四区": 4}.get(label)


def find_default_partition_xlsx() -> Path | None:
    candidates = sorted(BASE_DIR.glob("*.xlsx"))
    for path in candidates:
        if "分区" in path.name or "JCR" in path.name.upper():
            return path
    return candidates[0] if candidates else None


def import_partition_workbook(db: Session, xlsx_path: str | Path | None = None, force: bool = False) -> int:
    path = Path(xlsx_path) if xlsx_path else find_default_partition_xlsx()
    if not path or not path.exists():
        return 0
    if not force and db.query(JournalPartition).count() > 0:
        return 0

    if force:
        db.query(JournalPartition).delete()
        db.flush()

    workbook = load_workbook(path, read_only=True, data_only=True)
    imported = 0
    items_by_norm = {
        item.normalized_name: item
        for item in db.query(JournalPartition).all()
    }

    cas_sheet = _find_sheet(workbook, "中国科学院分区")
    if cas_sheet:
        for row in _iter_dict_rows(cas_sheet):
            name = row.get("期刊名称")
            normalized = normalize_journal_name(name)
            if not normalized:
                continue
            item = items_by_norm.get(normalized)
            if not item:
                item = JournalPartition(journal_name=str(name).strip(), normalized_name=normalized)
                db.add(item)
                items_by_norm[normalized] = item
                imported += 1
            item.cas_partition = partition_label(row.get("2025分区"))
            item.cas_partition_code = partition_code(row.get("2025分区"))
            item.is_top = is_yes(row.get("Top"))
            item.open_access = is_yes(row.get("Open Access"))
            item.imported_from = str(path)

    jcr_sheet = _find_sheet(workbook, "JCR")
    if jcr_sheet:
        for row in _iter_dict_rows(jcr_sheet):
            name = row.get("期刊名") or row.get("期刊名称")
            normalized = normalize_journal_name(name)
            if not normalized:
                continue
            item = items_by_norm.get(normalized)
            if not item:
                item = JournalPartition(
                    journal_name=str(name).strip(),
                    normalized_name=normalized,
                    cas_partition="未收录",
                    cas_partition_code=None,
                    imported_from=str(path),
                )
                db.add(item)
                items_by_norm[normalized] = item
                imported += 1
            item.issn = str(row.get("ISSN") or "").strip()
            item.eissn = str(row.get("eISSN") or "").strip()
            item.category = str(row.get("Category") or "").strip()
            item.impact_factor = _to_float(row.get("2024JIF"))
            item.source_type = source_type(item.category)
            item.imported_from = str(path)

    db.commit()
    return imported


def lookup_journal(db: Session, journal: str | None, doi: str | None = None) -> JournalPartition | None:
    candidates = [journal]
    normalized = normalize_journal_name(journal)

    for alias, canonical in ALIAS_PAIRS.items():
        if normalized == normalize_journal_name(alias):
            candidates.append(canonical)

    doi_value = str(doi or "").lower()
    for code, canonical in DOI_JOURNAL_CODES.items():
        if code in doi_value:
            candidates.append(canonical)

    for candidate in candidates:
        item = _lookup_exact(db, candidate)
        if item:
            return item

    if normalized and len(normalized) >= 6:
        keys = [row[0] for row in db.query(JournalPartition.normalized_name).all()]
        matches = difflib.get_close_matches(normalized, keys, n=1, cutoff=0.94)
        if matches:
            return db.query(JournalPartition).filter(JournalPartition.normalized_name == matches[0]).first()

    return None


def apply_partition_to_paper_data(db: Session, paper_data: dict[str, Any]) -> dict[str, Any]:
    match = lookup_journal(db, paper_data.get("journal"), paper_data.get("doi"))
    if not match:
        paper_data.setdefault("cas_partition", "未收录")
        paper_data.setdefault("is_top_journal", False)
        return paper_data

    paper_data["cas_partition"] = match.cas_partition or "未收录"
    paper_data["is_top_journal"] = bool(match.is_top)
    if not paper_data.get("journal"):
        paper_data["journal"] = match.journal_name
    for field, value in {
        "issn": match.issn,
        "eissn": match.eissn,
        "impact_factor": match.impact_factor,
        "source_type": match.source_type,
    }.items():
        if paper_data.get(field) in (None, ""):
            paper_data[field] = value
    return paper_data


def _lookup_exact(db: Session, value: str | None) -> JournalPartition | None:
    normalized = normalize_journal_name(value)
    if not normalized:
        return None
    return db.query(JournalPartition).filter(JournalPartition.normalized_name == normalized).first()


def _find_sheet(workbook, keyword: str):
    for sheet in workbook.worksheets:
        if keyword in sheet.title:
            return sheet
    return None


def _iter_dict_rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
    for row in rows:
        data = {headers[idx]: value for idx, value in enumerate(row) if idx < len(headers) and headers[idx]}
        if any(value not in (None, "") for value in data.values()):
            yield data


def _to_float(value: Any) -> float | None:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None
